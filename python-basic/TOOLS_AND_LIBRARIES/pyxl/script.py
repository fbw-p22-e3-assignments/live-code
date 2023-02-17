import openpyxl
#create venv
#activate
#pip install openpyxl
#file = openpyxl.load_workbook('mypyxl.xlsx')
wb = openpyxl.Workbook()
sheet_1 = wb.active  # opening a sheet called sheet_1 -- we don't need to create sheet since it is creates one sheet by default 
sheet_1['A1'] = 'Hello Excel'
sheet_1.title = 'Product'  # renaming sheet_1 to Product
sheet_2 = wb.create_sheet('Customer')   # we are creating a new sheet
sheet_2['C3'] = 'Hello Susan'  # making entry in the second sheet
file = wb.save('mypyxl.xlsx')   #created a file
# print((sheet_2.cell(row = 3, column = 3)))  #--> out : <Cell 'Customer'.C3>
print(sheet_2['C'])  # out --> (<Cell 'Customer'.C1>, <Cell 'Customer'.C2>, <Cell 'Customer'.C3>)
print(sheet_2['3'])  # out--> (<Cell 'Customer'.A3>, <Cell 'Customer'.B3>, <Cell 'Customer'.C3>)
print(type(sheet_2['3']))  # out --> <class 'tuple'>
for i in sheet_2['3']:
    print(i.value)   
"""None
   None
   Hello Susan"""
# for i in sheet_2['A1:E15']:
#     print(i)
for i in sheet_2['A1:E15']:
    for j in i:
        if j.value != None:
            print(j.value)
#for i in sheet_2.iter_rows(min_row = 2,  max_row = 7):  # if you don't specify min and max columns, it will print out all the columns that we have accessed before. in 34 we used E, so it will print upto E
    #print(i)
for i in sheet_2.iter_rows(min_row = 2,  max_row = 7, min_col = 1, max_col = 3, values_only=True):
    print(i)

print('#############')

for i in sheet_2.iter_cols(min_row = 2,  max_row = 7, min_col = 1, max_col = 3, values_only=True):
    print(i)
#close

# wb.close()
# #file.close()
# file = open('Path', 'r+')
# file.read()
# file.write('Hello word')
# file.close()

with openpyxl.load_workbook('mypyxl.xlsx') as wb:
    pass
